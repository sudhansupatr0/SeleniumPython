import time

import xlfile
from selenium import webdriver
import openpyxl

driver = webdriver.Chrome(executable_path="C:\Driver\chromedriver.exe")
driver.get("https://opensource-demo.orangehrmlive.com/index.php/auth/login")
driver.implicitly_wait(5)


UName = xlfile.readdata("C:\Driver\TestDataLoader.xlsx","test123",2,1)
Pword = xlfile.readdata("C:\Driver\TestDataLoader.xlsx","test123",2,2)
driver.find_element_by_id("txtUsername").send_keys(UName)
driver.find_element_by_id("txtPassword").send_keys(Pword)
driver.find_element_by_id("btnLogin").click()
time.sleep(5)
driver.save_screenshot("C:\Driver\screen1.jpg")















#data_path ="C:\Driver\TestDataLoader.xlsx"

#workbook = openpyxl.load_workbook(data_path)
#sheet = workbook.get_sheet_by_name("test123")
#rows = sheet.max_row
#cols = sheet.max_column
#print(rows)
#print(cols)

#for r in range(1,rows+1):
#    for c in range(1,cols+1):
#        print(sheet.cell(row=r,column=c).value)













#driver.get("http://testautomationpractice.blogspot.com/")
#driver.switch_to_frame(0)
#driver.find_element_by_id("RESULT_FileUpload-10").send_keys("C://Driver/TOSCA_CDS.pdf")












#driver.get("http://testautomationpractice.blogspot.com/")
#source_element = driver.find_element_by_id("draggable")
#target_element = driver.find_element_by_id("droppable")


#actions = ActionChains(driver)
#actions.drag_and_drop(source_element,target_element).perform()









#driver.get("https://opensource-demo.orangehrmlive.com/index.php/auth/login")
#driver.find_element_by_id("txtUsername").send_keys("Admin")
#driver.find_element_by_id("txtPassword").send_keys("admin123")
#driver.find_element_by_id("btnLogin").click()


#admin = driver.find_element_by_id("menu_admin_viewAdminModule")
#um = driver.find_element_by_id("menu_admin_UserManagement")
#users = driver.find_element_by_id("menu_admin_viewSystemUsers")
#actions = ActionChains(driver)
#actions.move_to_element(admin).move_to_element(um).move_to_element(users).click().perform()
















#time.sleep(5)
#driver.execute_script("window.scrollBy(0,500)","")
#time.sleep(5)
#Element = driver.find_element_by_xpath("//*[@id='HTML12']/h2")
#driver.execute_script("arguments[0].scrollIntoView();",Element)
#time.sleep(5)
#driver.execute_script("window.scrollBy(0,document.body.scrollHeight)")













#print(len(driver.find_elements_by_xpath("//table[@name='BookTable']/tbody/tr")))

#print(driver.find_element_by_xpath("//table[@name='BookTable']/tbody/tr[1]/th[1]").text)

#rows = len(driver.find_elements_by_xpath("//table[@name='BookTable']/tbody/tr"))
#cols = len(driver.find_elements_by_xpath("//table[@name='BookTable']/tbody/tr/th"))

#print(rows)
#print(cols)

#for r in range(2,rows+1):
#    for c in range(1,cols+1):
#       print( driver.find_element_by_xpath("//table[@name='BookTable']/tbody/tr["+str(r)+"]/td["+str(c)+"]").text)








#driver.get("http://testautomationpractice.blogspot.com/")
#driver.find_element_by_css_selector("button:nth-child(1)").click()
#driver.switch_to_alert().accept()
#driver.switch_to_window(1)






#driver.get("https://fs2.formsite.com/meherpavan/form2/index.html?1537702596407")

#driver.get("https://www.techlistic.com/p/selenium-practice-form.html")
#driver.find_element_by_css_selector("input[name='firstname']").send_keys("sudhansu")
#driver.find_element_by_css_selector("input[name='lastname']").send_keys("patro")
#driver.find_element_by_id("sex-0").click()
#driver.find_element_by_id("exp-6").click()
#driver.find_element_by_id("datepicker").send_keys("12/12/2020")
#driver.find_element_by_id("profession-1").click()
#driver.find_element_by_css_selector("input[value='Selenium Webdriver']").click()



#myele = driver.find_element_by_id("continents")
#myselect = Select(myele)

#myselect.select_by_index("2")
#print(len(myselect.options))
#all_options = myselect.options
#for option in all_options:
#    print(option.text)

#print(len(driver.find_elements_by_tag_name("a")))
#driver.find_element_by_link_text("Click here to Download File").click()

















#driver.find_element(By.CSS_SELECTOR("input[value='Radio-1']")).click()
#ena = driver.find_element_by_id("RESULT_RadioButton-7_0").is_enabled()
#print(ena)
#ena = driver.find_element_by_id("RESULT_RadioButton-7_0").is_selected()
#print(ena)

#element = driver.find_element_by_id("RESULT_CheckBox-8_0")
#driver.execute_script("arguments[0].click();", element)













#driver.find_element_by_css_selector("input[id='RESULT_TextField-1']").send_keys("sudhansu")
#print(len(driver.find_elements_by_class_name("text_field")))



















#wait = WebDriverWait(driver,10)

#wait.until(EC.element_to_be_clickable(By.CSS_SELECTOR("a button[class='btn btn-info']")),"RSDFASFA")

#print(driver.title)
#print(driver.current_url)
#driver.find_element_by_css_selector("a button[class='btn btn-info']").click()
#driver.quit()


















#driver.implicitly_wait(5)

#assert "Welcome: Mercury Tours" in driver.title

#ul = driver.find_element_by_name("userName")
#pwd= driver.find_element_by_name("password")
#sign_in = driver.find_element_by_name("login")

#if ul.is_enabled():
#    ul.send_keys("mercury")
#   pwd.send_keys("mercury")
#   sign_in.click()













#print(driver.title)

#driver.get("http://demo.automationtesting.in/Windows.html")
#print(driver.title)

#driver.back()
#print(driver.title)

#driver.forward()
#print(driver.title)


#print(driver.title)
#driver.maximize_window()
#driver.find_element_by_css_selector("a button[class='btn btn-info']").click()
#time.sleep(5);
#driver.close()
































#driver.quit()