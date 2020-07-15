import openpyxl


def getRowCount(file_path, sheetName):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(file_path, sheetName):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)


def readdata(file_path, sheetName , rowCount , ColumnCount):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowCount , column=ColumnCount).value

def writedata(file_path, sheetName , rowCount , ColumnCount , data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowCount , column=ColumnCount).value = data
    workbook.save(file_path)



